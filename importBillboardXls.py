import pandas as pd
from openpyxl import load_workbook
import os

# Define the path to the Excel file
file_path = 'import/billboard/Billboard Top 100 (1955-2023).xlsx'

# Get the directory of the source file
dir_path = os.path.dirname(file_path)

# Load the workbook
book = load_workbook(file_path)

# Get the sheet names (years)
sheet_names = book.sheetnames

# Iterate over the sheets
for sheet in sheet_names:
    # Read the sheet into a dataframe
    df = pd.read_excel(file_path, sheet_name=sheet, header=None, skiprows=2)
    
    # Rename the columns
    df.columns = ['Ranking', 'Title', 'Artist']

    # Strip all double quotes from the "Title" column
    df['Title'] = df['Title'].str.replace('"', '')

    # Add the "Played" column with a default value of 0 (not played yet)
    df['Played'] = 0
    
    # Reorder the columns
    df = df[['Ranking', 'Artist', 'Title', 'Played']]
    
    # Define the path to the CSV file
    tsv_path = os.path.join(dir_path, f'{sheet}.tsv')
    
    # Write the dataframe to a CSV file
    df.to_csv(tsv_path, sep='\t', index=True)